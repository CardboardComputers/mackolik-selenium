import numpy as np
from openpyxl.cell.cell import Cell
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from scipy.stats import poisson
from scipy.interpolate import interp1d
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
from copy import copy


def init_cell_style(cell: Cell) -> None:
    cell.font = Font(name='Consolas', color='FFFFFFFF')
    cell.alignment = Alignment(vertical='center')
    cell.fill = PatternFill(
        start_color='FF000000', end_color='FF000000', fill_type='solid')


def init_cross_dim(ws: Worksheet, index: int) -> None:
    ws.column_dimensions[get_column_letter(index)].width = 16
    ws.row_dimensions[index].height = 20


def clamp01(n: float) -> float:
    return 0.0 if n < 0 else 1.0 if n > 1 else n


def color_probability(p: float) -> PatternFill:
    COLOR_GRADIENT_INDICES = [0, 0.125, 0.25, 0.5, 1]
    COLOR_GRADIENT_STOPS_R = [0, 255, 0, 0, 0]
    COLOR_GRADIENT_STOPS_G = [0, 0, 0, 127, 127]
    COLOR_GRADIENT_STOPS_B = [0, 0, 255, 255, 0]
    COLOR_GRADIENT_R = interp1d(COLOR_GRADIENT_INDICES, COLOR_GRADIENT_STOPS_R)
    COLOR_GRADIENT_G = interp1d(COLOR_GRADIENT_INDICES, COLOR_GRADIENT_STOPS_G)
    COLOR_GRADIENT_B = interp1d(COLOR_GRADIENT_INDICES, COLOR_GRADIENT_STOPS_B)

    p = clamp01(p)
    r_hex = hex(round(float(COLOR_GRADIENT_R(p))))[2:].zfill(2)[0:2]
    g_hex = hex(round(float(COLOR_GRADIENT_G(p))))[2:].zfill(2)[0:2]
    b_hex = hex(round(float(COLOR_GRADIENT_B(p))))[2:].zfill(2)[0:2]
    
    color = 'FF{}{}{}'.format(r_hex, g_hex, b_hex)
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def highlight_result(cell: Cell) -> None:
    FILL_COLOR = 'FF7fBF7F'
    FONT_COLOR = 'FFFF0000'

    cell.fill = PatternFill(start_color=FILL_COLOR, end_color=FILL_COLOR, fill_type='solid')
    cell.font = Font(name='Consolas', color=FONT_COLOR, bold=True)


def select_cell(ws: Worksheet, row: int, column: int) -> Cell:
    target_cell = ws.cell(row, column)
    init_cell_style(target_cell)
    return target_cell


def generate_page(
    ws: Worksheet,
    ht_matches_played: int,
    ht_goals_scored: int,
    ht_goals_lost: int,
    vt_matches_played: int,
    vt_goals_scored: int,
    vt_goals_lost: int,
    league_matches_played: int,
    league_home_goals: int,
    league_visitor_goals: int,
    divide_by_ht_matches_played: bool = False
) -> None:

    THICK_SIDE = Side(style='thick', color='FFFFFFFF')
    DOUBLE_SIDE = Side(style='double', color='FFFFFFFF')

    # TODO: figure out what to do if divisors are 0
    if ht_matches_played == 0:
        ht_matches_played = 1
    if vt_matches_played == 0:
        vt_matches_played = 1
    if league_matches_played == 0:
        league_matches_played = 1

    ht_mean_goals_scored = ht_goals_scored / ht_matches_played
    ht_mean_goals_lost = ht_goals_lost / ht_matches_played
    vt_mean_goals_scored = vt_goals_scored / vt_matches_played
    vt_mean_goals_lost = vt_goals_lost / vt_matches_played
    league_mean_home_goals = league_home_goals / league_matches_played
    league_mean_visitor_goals = league_visitor_goals / league_matches_played

    # TODO: figure out what to do if divisors are 0
    if league_mean_home_goals == 0:
        league_mean_home_goals = 1
    if league_mean_visitor_goals == 0:
        league_mean_visitor_goals = 1

    ht_power_offense = ht_mean_goals_scored / league_mean_home_goals
    ht_power_defense = ht_mean_goals_lost / league_mean_visitor_goals
    vt_power_offense = vt_mean_goals_scored / league_mean_visitor_goals
    vt_power_defense = vt_mean_goals_lost / league_mean_home_goals

    ht_adj_mean_goals = ht_power_offense * vt_power_defense * league_mean_home_goals
    vt_adj_mean_goals = vt_power_offense * ht_power_defense * league_mean_visitor_goals

    working_cell = select_cell(ws, 1, 1)
    working_cell.value = 'Gol sayısı'
    working_cell.border = Border(right=THICK_SIDE, bottom=THICK_SIDE)

    working_cell = select_cell(ws, 2, 2)
    working_cell.value = 'Olasılık'
    working_cell.border = Border(right=DOUBLE_SIDE, bottom=DOUBLE_SIDE)

    working_cell = select_cell(ws, 1, 2)
    working_cell.value = 'Ev Sahibi Takım'
    working_cell.border = Border(right=DOUBLE_SIDE, bottom=THICK_SIDE)

    working_cell = select_cell(ws, 2, 1)
    working_cell.value = 'Konuk Takım'
    working_cell.border = Border(right=THICK_SIDE, bottom=DOUBLE_SIDE)
    
    init_cross_dim(ws, 1)
    init_cross_dim(ws, 2)

    # calculate probabilities for each team to score a given number of goals
    for i in range(8):
        # axis along top
        working_cell = select_cell(ws, 1, 3+i)
        working_cell.value = str(i)
        working_cell.border = Border(bottom=THICK_SIDE)
        # axis along left
        working_cell = select_cell(ws, 3+i, 1)
        working_cell.value = str(i)
        working_cell.border = Border(right=THICK_SIDE)

        # home team goal probabilities
        ht_cell = select_cell(ws, 2, 3+i)
        ht_value = poisson.pmf(k=i, mu=ht_adj_mean_goals)
        ht_cell.value = '{:.9f}'.format(ht_value)
        ht_cell.fill = color_probability(ht_value)
        ht_cell.border = Border(bottom=DOUBLE_SIDE)
        # visitor team goal probabilities
        vt_cell = select_cell(ws, 3+i, 2)
        vt_value = poisson.pmf(k=i, mu=vt_adj_mean_goals)
        vt_cell.value = '{:.9f}'.format(vt_value)
        vt_cell.fill = color_probability(vt_value)
        vt_cell.border = Border(right=DOUBLE_SIDE)

        init_cross_dim(ws, 3+i)

    # calculate probabilities for all combinations of goals
    for i in range(8):
        ht_cell = ws.cell(2, 3+i)
        for j in range(8):
            vt_cell = ws.cell(3+j, 2)
        
            working_cell = select_cell(ws, 3+j, 3+i)
            working_value = float(ht_cell.value) * float(vt_cell.value)
            working_cell.value = '{:.9f}'.format(working_value)
            working_cell.fill = color_probability(working_value)

    # set up table for summed match score probabilities
    for r in range(11, 17):
        ws.row_dimensions[r].height = 20

    working_cell = select_cell(ws, 12, 1)
    working_cell.value = 'n'
    working_cell.border = Border(right=THICK_SIDE)
    working_cell = select_cell(ws, 13, 1)
    working_cell.value = 'n alt bitme'
    working_cell.border = Border(right=THICK_SIDE)
    working_cell = select_cell(ws, 14, 1)
    working_cell.value = '1/alt/ev oyna'
    working_cell.border = Border(right=THICK_SIDE)
    working_cell = select_cell(ws, 15, 1)
    working_cell.value = 'n üst bitme'
    working_cell.border = Border(right=THICK_SIDE)
    working_cell = select_cell(ws, 16, 1)
    working_cell.value = '1/üst/ev oyna'
    working_cell.border = Border(right=THICK_SIDE)
    for i in range(7):
        working_cell = select_cell(ws, 12, 2+i)
        working_cell.value = '{},5'.format(i)

    # P(<0,5) === P(0:0)
    working_cell = select_cell(ws, 13, 2)
    working_cell.value = ws.cell(3, 3).value
    working_cell.fill = copy(ws.cell(3, 3).fill)
    # P(<1,5) etc.
    for score_sum in range(1, 7):
        working_cell = select_cell(ws, 13, 2+score_sum)
        # start with probability for score_sum - 1
        working_value = float(ws.cell(13, 1+score_sum).value)
        # add all the probabilities for match scores that add to score_sum
        for i in range(score_sum + 1):
            j = score_sum - i
            working_value += float(ws.cell(3+i, 3+j).value)
        # set value and color
        working_cell.value = '{:.9f}'.format(working_value)
        working_cell.fill = color_probability(working_value)

    # all other values
    for i in range(7):
        # 1 / P(<n,5) / home team games played
        working_cell = select_cell(ws, 14, 2+i)
        if divide_by_ht_matches_played:
            working_value = np.reciprocal(float(ws.cell(13, 2+i).value) * ht_matches_played)
        else:
            working_value = np.reciprocal(float(ws.cell(13, 2+i).value))
        working_cell.value = '{:.9f}'.format(working_value)
        highlight_result(working_cell)
        # P(>n,5)
        working_cell = select_cell(ws, 15, 2+i)
        working_value = 1 - float(ws.cell(13, 2+i).value)
        working_cell.value = '{:.9f}'.format(working_value)
        working_cell.fill = color_probability(working_value)
        # 1 / P(>n,5) / home team games played
        working_cell = select_cell(ws, 16, 2+i)
        if divide_by_ht_matches_played:
            working_value = np.reciprocal(float(ws.cell(15, 2+i).value) * ht_matches_played)
        else:
            working_value = np.reciprocal(float(ws.cell(15, 2+i).value))
        working_cell.value = '{:.9f}'.format(working_value)
        highlight_result(working_cell)

    # create table of initial values
    for r in range(17, 21):
        ws.row_dimensions[r].height = 20
    ws.merge_cells(range_string='A18:C18')
    ws.merge_cells(range_string='D18:F18')
    ws.merge_cells(range_string='G18:I18')
    working_cell = select_cell(ws, 18, 1)
    working_cell.alignment = Alignment('center', 'center')
    working_cell.value = 'Ev Sahibi Takım'
    working_cell.border = Border(left=THICK_SIDE, right=THICK_SIDE, bottom=THICK_SIDE)
    working_cell = select_cell(ws, 18, 4)
    working_cell.alignment = Alignment('center', 'center')
    working_cell.value = 'Konuk Takım'
    working_cell.border = Border(left=THICK_SIDE, right=THICK_SIDE, bottom=THICK_SIDE)
    working_cell = select_cell(ws, 18, 7)
    working_cell.alignment = Alignment('center', 'center')
    working_cell.value = 'Lig Ortalaması'
    working_cell.border = Border(left=THICK_SIDE, right=THICK_SIDE, bottom=THICK_SIDE)
    for i in range(1, 10):
        working_cell = select_cell(ws, 19, i)
        working_cell.border = Border(left=THICK_SIDE, right=THICK_SIDE, bottom=THICK_SIDE)
        working_cell = select_cell(ws, 20, i)
        working_cell.border = Border(left=THICK_SIDE, right=THICK_SIDE, bottom=THICK_SIDE)
    
    ws.cell(19, 1).value = 'Oynadığı maç sayısı'
    ws.cell(19, 2).value = 'Attığı gol'
    ws.cell(19, 3).value = 'Yediği gol'
    ws.cell(19, 4).value = 'Oynadığı maç sayısı'
    ws.cell(19, 5).value = 'Attığı gol'
    ws.cell(19, 6).value = 'Yediği gol'
    ws.cell(19, 7).value = 'Oynanan maç sayısı'
    ws.cell(19, 8).value = 'Ev sahibi gol'
    ws.cell(19, 9).value = 'Misafir gol'

    ws.cell(20, 1).value = str(ht_matches_played)
    ws.cell(20, 2).value = str(ht_goals_scored)
    ws.cell(20, 3).value = str(ht_goals_lost)
    ws.cell(20, 4).value = str(vt_matches_played)
    ws.cell(20, 5).value = str(vt_goals_scored)
    ws.cell(20, 6).value = str(vt_goals_lost)
    ws.cell(20, 7).value = str(league_matches_played)
    ws.cell(20, 8).value = str(league_home_goals)
    ws.cell(20, 9).value = str(league_visitor_goals)


    # list format at bottom of page
    select_cell(ws, 22, 1).value = 'Gol sayısı'
    select_cell(ws, 22, 2).value = 'Ev Sahibi Takım'
    select_cell(ws, 22, 3).value = 'Konuk Takım'
    for i in range(8):
        value_col = 3+i
        ws.row_dimensions[21 + 9 * i].height = 20

        select_cell(ws, 23 + i, 1).value = str(i)
        select_cell(ws, 23 + i, 2).value = ws.cell(2, value_col).value
        select_cell(ws, 23 + i, 3).value = ws.cell(value_col, 2).value
        for j in range(8):
            value_row = 3+j
        
            offset_row = j + 9 * i
            ws.row_dimensions[22 + offset_row].height = 20
            working_cell = select_cell(ws, 22 + offset_row, 5)
            working_cell.value = '{} - {}'.format(i, j)
            working_cell = select_cell(ws, 22 + offset_row, 6)
            working_cell.value = ws.cell(value_row, value_col).value



def write_spreadsheet(
    filename,
    # inputs for normal page
    ht_total_matches_played: int,
    ht_total_goals_scored: int,
    ht_total_goals_lost: int,
    at_total_matches_played: int,
    at_total_goals_scored: int,
    at_total_goals_lost: int,
    league_home_matches_played: int,
    # inputs for double page
    ht_home_matches_played: int,
    ht_home_goals_scored: int,
    ht_home_goals_lost: int,
    at_away_matches_played: int,
    at_away_goals_scored: int,
    at_away_goals_lost: int,
    league_total_matches_played: int,
    # common inputs
    league_home_goals: int,
    league_away_goals: int
) -> None:

    workbook: Workbook = Workbook()
    ws = workbook.active
    ws.title = 'Normal toplam maç sayısı'

    generate_page(
        ws,
        ht_total_matches_played,
        ht_total_goals_scored,
        ht_total_goals_lost,
        at_total_matches_played,
        at_total_goals_scored,
        at_total_goals_lost,
        league_home_matches_played,
        league_home_goals,
        league_away_goals
    )

    workbook.create_sheet('Çift toplam maç sayısı')
    ws = workbook['Çift toplam maç sayısı']

    generate_page(
        ws,
        ht_home_matches_played,
        ht_home_goals_scored,
        ht_home_goals_lost,
        at_away_matches_played,
        at_away_goals_scored,
        at_away_goals_lost,
        league_total_matches_played,
        league_home_goals,
        league_away_goals,
        True
    )

    workbook.save(filename)

